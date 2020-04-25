#!/usr/bin/env python3
''' Excel操作サンプル
'''

import sys, os, datetime, openpyxl
from openpyxl.styles import Font

# 別ディレクトリのため、パス追加
script_dir = os.path.abspath(__file__)
base_dir = os.path.join(os.path.dirname(script_dir), '..')
sys.path.append(os.path.join(base_dir, 'com'))

from console import *
log('#============================')
log('# Excel')
log('#============================')

# ファイル読込
home_dir = os.path.expanduser('~')
excel_dir = os.path.join(home_dir, 'work', 'excel', 'automatestuff-ja', 'ch12')
excel_path = os.path.join(excel_dir, 'example.xlsx')

# workbook  
wb = openpyxl.load_workbook(excel_path)
log('WORK BOOK', type(wb))
log_add_line(1)

# シート名取得
# deprecated
# sheet_names = wb.get_sheet_names()
sheet_names = wb.sheetnames
log('SHEET NAME', sheet_names)
log_add_line(1)

# sheetオブジェクト取得
# deprecated
# sheet = wb.get_sheet_by_name('Sheet3')
sheet = wb['Sheet3']
log('SHEET', type(sheet))
log('SHEET TITLE', sheet.title)
log_add_line(1)

# active sheet
active_sheet = wb.active
log('ACTIVE SHEET', active_sheet.title)
log_add_line(1)

# cell取得
cell_a1 = active_sheet['A1']
log('CELL A1', cell_a1.value)
log('CELL ROW', cell_a1.row)
log('CELL COLUMN', cell_a1.column)
log('CELL COORDINATE', cell_a1.coordinate)
log_add_line(1)

# sheet1の表全体読込
log('> 表全体読込 - for')
active_sheet.max_row
active_sheet.max_column
for row in range(1, active_sheet.max_row + 1):
    for column in range(1, active_sheet.max_column + 1):
        log('{0}行 {1}列'.format(row, column), active_sheet.cell(row, column).value)
    log('--- END OF ROW')
log_add_line(1)

log('> 表全体読込 - tupple')
rows = tuple(active_sheet['A1':'C3'])
for row in rows:
    for cell in row:
        log('{0}行 {1}列'.format(cell.row, cell.column), cell.value)
    log('--- END OF ROW')
log_add_line()


# シート名変更の作成
log('> CREATE NEW SHEET')
log(' - rename sheet')
active_sheet.title = 'RENAME_{0}'.format(active_sheet.title)
# シート追加
log(' - add sheet')
now = datetime.datetime.now()
formatted_now = now.strftime('%Y%m%d_%H%M%S')
new_sheet_title = 'NEW_{0}'.format(formatted_now)
new_sheet = wb.create_sheet()
new_sheet.title = new_sheet_title
new_sheet.cell(1, 1).value = 'NEW VALUE'
log('NEW SHEET CELL', new_sheet.cell(1, 1).value)
# シート削除
log(' - remove sheet')
#wb.remove_sheet(wb['Sheet2'])
wb.remove(wb['Sheet2'])
wb.remove(wb['Sheet3'])
# 別ファイルで保存
wb.save(os.path.join(excel_dir, 'example_new.xlsx'))
log_add_line()


log('> CREATE NEW BOOK')
# セルのフォントスタイル設定
wb = openpyxl.Workbook()
sheet = wb.create_sheet()
sheet.title = 'style_sample'
# Fontには設定可能な属性がいくつか存在する
# name : フォント
# size : 文字サイズ
# bold : 太文字（boolean）
# italic : 斜文字（boolean）
italic24_font = Font(size=24, italic=True, color='FF0000')
sheet.cell(1, 1).font = italic24_font
sheet.cell(1, 1).value = 'Hello World'

# 数式
sheet = wb.create_sheet()
sheet.title = 'calculation'
sheet.cell(1, 1).value = 100
sheet.cell(2, 1).value = 200
sheet.cell(3, 1).value = 900
sheet.cell(4, 1).value = '=sum(A1:A3)'
# 保存
# wb.save(os.path.join(excel_dir, 'sample.xlsx'))

# 計算結果の値がほしい場合は、data_only属性を指定取得する必要がある
# workbookは計算結果か式かいずれかしか取得できない
data_wb = openpyxl.load_workbook(os.path.join(excel_dir, 'sample.xlsx'), data_only=True)
data_sheet = data_wb['calculation']
log(type(data_sheet))
# 何故かNoneが返される。。。想定としては1200
log('SUM', data_sheet.cell(4, 1).value)
